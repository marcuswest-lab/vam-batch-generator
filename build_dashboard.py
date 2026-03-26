import pandas as pd
import openpyxl
import json
import re
from collections import defaultdict

f_perf = '[VAM] - Creative & Copy Tracking Sheet - 2 ROAS, $250 cost per connected, CAC _ $750.xlsx'

# Example/test ad names to exclude
EXCLUDE_NAMES = {'SC1_9t5 | SV1_beach1', 'CHK1_9t5 | CB1_beach1'}

# --- Extract preview/file links from tracker tabs (including hyperlinks) ---
def extract_links_from_tracker(sheet_name, name_header='Creative Name', new_name_header='New Name'):
    wb = openpyxl.load_workbook(f_perf, data_only=True)
    ws = wb[sheet_name]
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    name_col = headers.get(name_header)
    new_name_col = headers.get(new_name_header)
    file_link_col = headers.get('File Link')
    perf_col = headers.get('Performance')
    new_perf_col = headers.get('New Performance')

    link_map = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cname = row[name_col-1].value if name_col else None
        nname = row[new_name_col-1].value if new_name_col and new_name_col <= len(row) else None

        best_link = None
        if file_link_col:
            cell = row[file_link_col-1]
            val = str(cell.value or '')
            if 'http' in val.lower(): best_link = val
            elif cell.hyperlink and cell.hyperlink.target and 'http' in cell.hyperlink.target.lower():
                best_link = cell.hyperlink.target

        if not best_link:
            for pcol in [new_perf_col, perf_col]:
                if pcol:
                    cell = row[pcol-1]
                    if cell and cell.hyperlink and cell.hyperlink.target and 'http' in cell.hyperlink.target.lower():
                        best_link = cell.hyperlink.target
                        break

        if not best_link: continue
        if nname and str(nname) != 'nan': link_map[str(nname)] = best_link
        if cname and str(cname) != 'nan': link_map[str(cname)] = best_link

    wb.close()
    return link_map

def extract_links_from_new_tracker(file_path, sheet_name, name_header):
    """Extract links from the new Creative Tracker (File 1) using Folder Link column."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    name_col = headers.get(name_header)
    folder_col = headers.get('Folder Link')
    if not name_col or not folder_col:
        wb.close()
        return {}

    link_map = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        name_cell = row[name_col-1]
        folder_cell = row[folder_col-1]

        name = str(name_cell.value or '') if name_cell else ''
        if not name or name == 'nan': continue

        link = None
        if folder_cell:
            val = str(folder_cell.value or '')
            if 'http' in val.lower(): link = val
            elif folder_cell.hyperlink and folder_cell.hyperlink.target and 'http' in folder_cell.hyperlink.target.lower():
                link = folder_cell.hyperlink.target

        if link:
            link_map[name] = link

    wb.close()
    return link_map

f_tracker = '[Value Added Moving] - Creative Tracker.xlsx'

print('Extracting links from tracker tabs...')
# Old tracker (Creative & Copy Tracking Sheet)
STATIC_LINKS = extract_links_from_tracker('Static Ad Tracker', 'Creative Name', 'New Name')
VIDEO_LINKS = extract_links_from_tracker('Video Ad Tracker', 'Creative Name', 'New Name')
COPY_LINKS = extract_links_from_tracker('Copy Tracker', 'Creative Name', None)
print(f'  Old tracker — Static: {len(STATIC_LINKS)}, Video: {len(VIDEO_LINKS)}, Copy: {len(COPY_LINKS)}')

# New tracker (Creative Tracker) — has links for new-style ad names
NEW_STATIC_LINKS = extract_links_from_new_tracker(f_tracker, 'Static Creative Tracker', 'Static Creative Name')
NEW_VIDEO_LINKS = extract_links_from_new_tracker(f_tracker, 'Video Creative Tracker', 'Video Creative Name')
print(f'  New tracker — Static: {len(NEW_STATIC_LINKS)}, Video: {len(NEW_VIDEO_LINKS)}')

# Merge: new tracker links take priority (more specific file links)
STATIC_LINKS.update(NEW_STATIC_LINKS)
VIDEO_LINKS.update(NEW_VIDEO_LINKS)
print(f'  Merged — Static: {len(STATIC_LINKS)}, Video: {len(VIDEO_LINKS)}, Copy: {len(COPY_LINKS)}')

def parse_dollar(v):
    if pd.isna(v): return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).replace('$','').replace(',','')
    try: return float(s)
    except: return None

def clean_pct(v):
    if pd.isna(v) or str(v).strip() in ('', '#DIV/0!', '#REF!', 'nan'): return None
    if isinstance(v, (int, float)): return round(v * 100, 2) if abs(v) < 1 else round(v, 2)
    s = str(v).replace('%','').strip()
    try: return round(float(s), 2)
    except: return None

def clean_num(v):
    if pd.isna(v) or str(v).strip() in ('', '#DIV/0!', '#REF!', 'nan'): return None
    try: return float(v)
    except: return None

def process_sheet(sheet_name, is_video=False, link_map=None):
    df = pd.read_excel(f_perf, sheet_name=sheet_name)
    df['Cost_num'] = df['Cost'].apply(parse_dollar)
    df = df[df['Cost_num'] > 0].copy()
    if link_map is None: link_map = {}

    rows = []
    for _, r in df.iterrows():
        ad_name = str(r.get('Ad Name', ''))
        if not ad_name or ad_name == 'nan': continue
        if ad_name in EXCLUDE_NAMES: continue

        cost = parse_dollar(r.get('Cost'))
        leads = clean_num(r.get('Leads'))
        cpl = parse_dollar(r.get('Cost per Lead'))
        new_leads = clean_num(r.get('New Leads'))
        cpnl = parse_dollar(r.get('Cost per New Lead'))
        incoming = clean_num(r.get('Stage: Incoming Call'))
        cost_incoming = parse_dollar(r.get('cost per incoming call'))
        outbound = clean_num(r.get('api-outbound-call-outbound / Calls'))
        cost_outbound = parse_dollar(r.get('Cost per Outbound Call'))
        inbound_connected = clean_num(r.get('Stage: Inbound Connected Call'))
        cost_inbound_connected = parse_dollar(r.get('cost per inbound connected'))
        outbound_connected = clean_num(r.get('Outbound connected calls'))
        cost_outbound_connected = parse_dollar(r.get('cost per outbound connected'))
        connected = clean_num(r.get('Connected (Total)'))
        cost_connected = parse_dollar(r.get('cost per connected (total)'))
        conn_close = clean_pct(r.get('connected close rate (total)'))
        sales = clean_num(r.get('Sales'))
        cps = parse_dollar(r.get('Cost per Sale'))
        revenue = parse_dollar(r.get('Revenue'))
        roas = clean_num(r.get('ROAS'))
        profit = parse_dollar(r.get('Profit'))
        ctr_all = clean_pct(r.get('CTR (all)'))
        ctr_link = clean_pct(r.get('CTR (link)'))
        cpm = parse_dollar(r.get('CPM'))
        impressions = clean_num(r.get('Impressions'))
        status = str(r.get('Status', ''))

        row = {
            'name': ad_name, 'status': status, 'cost': cost, 'leads': leads,
            'cpl': cpl, 'newLeads': new_leads, 'cpnl': cpnl,
            'incoming': incoming, 'costIncoming': cost_incoming,
            'outbound': outbound, 'costOutbound': cost_outbound,
            'inboundConnected': inbound_connected, 'costInboundConnected': cost_inbound_connected,
            'outboundConnected': outbound_connected, 'costOutboundConnected': cost_outbound_connected,
            'connected': connected, 'costConnected': cost_connected,
            'connClose': conn_close, 'sales': sales, 'cps': cps,
            'revenue': revenue, 'roas': roas, 'profit': profit,
            'ctrAll': ctr_all, 'ctrLink': ctr_link, 'cpm': cpm,
            'impressions': impressions,
            'link': link_map.get(ad_name),
        }

        if is_video:
            row['hookRate'] = clean_pct(r.get('Hook Rate'))
            row['holdRate'] = clean_pct(r.get('Hold Rate'))

        rows.append(row)

    rows.sort(key=lambda x: -(x['roas'] if x['roas'] is not None else -999))
    return rows

# --- Extract concept/angle name from ad name ---
def normalize_old_concept(raw):
    """Strip version/holiday/CTA/NY suffixes to get core concept name.
    e.g. CheckThisHolidayV3 -> CheckThis, LongDistance1597HolidayV2CTA1 -> LongDistance1597,
    CityPricesV4 -> CityPrices, ChooseDateLaterV2 -> ChooseDateLater,
    ArrivedFlawlessV1 -> ArrivedFlawless, CoastToCoast-V2 -> CoastToCoast"""
    # Strip CTA suffixes first (CTA1, CTA2)
    raw = re.sub(r'CTA\d+$', '', raw)
    # Strip Holiday/NY variants (HolidayV1, HolidayV2, NYV1, NYV2)
    raw = re.sub(r'Holiday(V\d+)?$', '', raw)
    raw = re.sub(r'NY(V\d+)?$', '', raw)
    # Strip trailing version numbers (V1, V2, -V1, -V2) but NOT if they're part of the core name
    # Only strip if preceded by a lowercase letter or digit (not part of a word like "V1_something")
    raw = re.sub(r'-?V\d+$', '', raw)
    return raw.strip('-').strip()

def extract_static_concept(name):
    """Extract concept name for grouping statics."""
    if '|' in name:
        sc_part = name.split('|')[0].strip()
        m = re.match(r'SC\d+_(.+)', sc_part)
        if m:
            return normalize_old_concept(m.group(1))
        return sc_part
    elif ' - ' in name:
        return name.split(' - ')[0].strip()
    return name

def extract_video_concept(name):
    """Extract concept name for grouping videos."""
    if '|' in name:
        parts = [p.strip() for p in name.split('|')]
        vv = [p for p in parts if p.startswith('VV')]
        if vv:
            m = re.match(r'VV\d+_(.+)', vv[0])
            raw = m.group(1) if m else vv[0]
            return normalize_old_concept(raw)
        return parts[0]
    elif ' - ' in name:
        return name.split(' - ')[0].strip()
    return name

def extract_copy_concept(name):
    """Extract concept name for grouping copy."""
    if '|' in name:
        # CHK1_SaveUpTo50 | CB1_Savings -> group by CHK (hook concept)
        chk_part = name.split('|')[0].strip()
        m = re.match(r'CHK\d+_(.+)', chk_part)
        if m:
            return normalize_old_concept(m.group(1))
        return chk_part
    elif ' - ' in name:
        return name.split(' - ')[0].strip()
    return name

def rollup(rows, concept_fn, is_video=False):
    """Roll up variation-level rows into concept-level aggregates."""
    groups = defaultdict(list)
    for r in rows:
        concept = concept_fn(r['name'])
        groups[concept].append(r)

    rolled = []
    for concept, variations in groups.items():
        cost = sum(v['cost'] or 0 for v in variations)
        leads = sum(v['leads'] or 0 for v in variations)
        new_leads = sum(v.get('newLeads') or 0 for v in variations)
        incoming = sum(v['incoming'] or 0 for v in variations)
        outbound = sum(v.get('outbound') or 0 for v in variations)
        inbound_connected = sum(v.get('inboundConnected') or 0 for v in variations)
        outbound_connected = sum(v.get('outboundConnected') or 0 for v in variations)
        connected = sum(v['connected'] or 0 for v in variations)
        sales = sum(v['sales'] or 0 for v in variations)
        revenue = sum(v['revenue'] or 0 for v in variations)
        profit = sum(v['profit'] or 0 for v in variations)
        impressions = sum(v['impressions'] or 0 for v in variations)

        row = {
            'name': concept,
            'status': '',
            'variations': len(variations),
            'cost': cost if cost > 0 else None,
            'leads': leads if leads > 0 else None,
            'cpl': (cost / leads) if leads > 0 else None,
            'newLeads': new_leads if new_leads > 0 else None,
            'cpnl': (cost / new_leads) if new_leads > 0 else None,
            'incoming': incoming if incoming > 0 else None,
            'costIncoming': (cost / incoming) if incoming > 0 else None,
            'outbound': outbound if outbound > 0 else None,
            'costOutbound': (cost / outbound) if outbound > 0 else None,
            'inboundConnected': inbound_connected if inbound_connected > 0 else None,
            'costInboundConnected': (cost / inbound_connected) if inbound_connected > 0 else None,
            'outboundConnected': outbound_connected if outbound_connected > 0 else None,
            'costOutboundConnected': (cost / outbound_connected) if outbound_connected > 0 else None,
            'connected': connected if connected > 0 else None,
            'costConnected': (cost / connected) if connected > 0 else None,
            'connClose': (sales / connected * 100) if connected > 0 and sales > 0 else None,
            'sales': sales if sales > 0 else None,
            'cps': (cost / sales) if sales > 0 else None,
            'revenue': revenue if revenue > 0 else None,
            'roas': (revenue / cost) if cost > 0 and revenue > 0 else None,
            'profit': profit,
            'ctrAll': None,
            'ctrLink': None,
            'cpm': (cost / impressions * 1000) if impressions > 0 else None,
            'impressions': impressions if impressions > 0 else None,
        }

        if is_video:
            hook_total = sum((v.get('hookRate') or 0) * (v.get('impressions') or 0) for v in variations)
            hold_total = sum((v.get('holdRate') or 0) * (v.get('impressions') or 0) for v in variations)
            total_impr = sum(v.get('impressions') or 0 for v in variations)
            row['hookRate'] = round(hook_total / total_impr, 1) if total_impr > 0 and hook_total > 0 else None
            row['holdRate'] = round(hold_total / total_impr, 1) if total_impr > 0 and hold_total > 0 else None

        # Pick link from top-spending variation that has a link
        linked = [v for v in variations if v.get('link')]
        if linked:
            top_var = max(linked, key=lambda v: v.get('cost') or 0)
            row['link'] = top_var['link']
        else:
            row['link'] = None

        rolled.append(row)

    rolled.sort(key=lambda x: -(x['roas'] if x['roas'] is not None else -999))
    return rolled

# --- Process all sheets ---
static_data = process_sheet('Static Ad Performance', link_map=STATIC_LINKS)
video_data = process_sheet('Video Ad Performance', is_video=True, link_map=VIDEO_LINKS)
copy_data = process_sheet('Copy Performance', link_map=COPY_LINKS)

# Deduplicate static data
seen = set()
deduped_static = []
for r in static_data:
    if r['name'] not in seen:
        seen.add(r['name'])
        deduped_static.append(r)
static_data = deduped_static

# Also exclude example rows from video/copy (VHK1_9t5 patterns)
video_data = [r for r in video_data if '9t5' not in r['name']]
copy_data = [r for r in copy_data if '9t5' not in r['name']]

# Build roll-ups
static_concepts = rollup(static_data, extract_static_concept)
video_concepts = rollup(video_data, extract_video_concept, is_video=True)
copy_concepts = rollup(copy_data, extract_copy_concept)

def to_js(rows):
    return json.dumps(rows)

static_json = to_js(static_data)
video_json = to_js(video_data)
copy_json = to_js(copy_data)
static_concepts_json = to_js(static_concepts)
video_concepts_json = to_js(video_concepts)
copy_concepts_json = to_js(copy_concepts)

dashboard_html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>VAM Creative Performance Dashboard</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:#0C0F14; color:#E2E8F0; font-family:'DM Sans',sans-serif; }}
.container {{ max-width:1700px; margin:0 auto; padding:24px 32px; }}
header {{ text-align:center; margin-bottom:24px; }}
header h1 {{ font-size:28px; font-weight:700; color:#fff; margin-bottom:4px; }}
header p {{ font-size:14px; color:#94A3B8; }}
.tabs {{ display:flex; gap:0; margin-bottom:0; background:#151922; border-radius:10px 10px 0 0; border:1px solid #1E293B; border-bottom:none; overflow:hidden; }}
.tab {{ padding:12px 24px; cursor:pointer; font-size:13px; font-weight:600; color:#64748B; transition:all 0.2s; border-right:1px solid #1E293B; }}
.tab:last-child {{ border-right:none; }}
.tab:hover {{ color:#CBD5E1; background:#1a2030; }}
.tab.active {{ color:#fff; background:#1E293B; }}
.tab .count {{ font-size:11px; color:#475569; margin-left:6px; font-weight:400; }}
.tab.active .count {{ color:#94A3B8; }}
.table-wrap {{ background:#151922; border:1px solid #1E293B; border-radius:0 0 10px 10px; overflow-x:auto; max-height:80vh; overflow-y:auto; }}
table {{ border-collapse:collapse; width:100%; min-width:2200px; }}
thead {{ position:sticky; top:0; z-index:10; }}
th {{ background:#1E293B; padding:10px 12px; text-align:left; font-size:11px; font-weight:600; color:#94A3B8; text-transform:uppercase; letter-spacing:0.3px; white-space:nowrap; border-bottom:2px solid #334155; cursor:pointer; user-select:none; }}
th:hover {{ color:#E2E8F0; }}
th .arrow {{ font-size:10px; margin-left:3px; opacity:0.4; }}
th.sorted .arrow {{ opacity:1; color:#818CF8; }}
td {{ padding:8px 12px; font-size:12px; border-bottom:1px solid #1E293B; white-space:nowrap; font-family:'JetBrains Mono',monospace; color:#CBD5E1; }}
td.name-col {{ font-family:'DM Sans',sans-serif; font-weight:500; position:sticky; left:0; background:#151922; z-index:5; min-width:320px; max-width:420px; overflow:hidden; text-overflow:ellipsis; }}
td.num-col {{ text-align:right; }}
td.idx-col {{ text-align:center; color:#475569; font-size:11px; position:sticky; left:0; background:#151922; z-index:5; width:36px; }}
tr:hover td {{ background:#1a2030; }}
tr:hover td.name-col, tr:hover td.idx-col {{ background:#1a2030; }}
.badge {{ display:inline-block; padding:2px 8px; border-radius:4px; font-size:11px; font-weight:600; font-family:'JetBrains Mono',monospace; }}
.badge.wk {{ background:rgba(52,211,153,0.15); color:#34D399; }}
.badge.of {{ background:rgba(251,191,36,0.15); color:#FBBF24; }}
.badge.ak {{ background:rgba(248,113,113,0.15); color:#F87171; }}
.badge.nd {{ background:rgba(100,116,139,0.1); color:#475569; }}
.badge.status-winner {{ background:rgba(52,211,153,0.15); color:#34D399; font-family:'DM Sans',sans-serif; font-size:10px; }}
.badge.status-loser {{ background:rgba(248,113,113,0.15); color:#F87171; font-family:'DM Sans',sans-serif; font-size:10px; }}
.badge.status-testing {{ background:rgba(96,165,250,0.15); color:#60A5FA; font-family:'DM Sans',sans-serif; font-size:10px; }}
.badge.status-fence {{ background:rgba(251,191,36,0.15); color:#FBBF24; font-family:'DM Sans',sans-serif; font-size:10px; }}
.badge.status-other {{ background:rgba(100,116,139,0.1); color:#94A3B8; font-family:'DM Sans',sans-serif; font-size:10px; }}
.legend {{ display:flex; gap:24px; justify-content:center; margin:0 0 16px; flex-wrap:wrap; font-size:12px; color:#94A3B8; }}
.legend-item {{ display:flex; align-items:center; gap:6px; }}
.legend-dot {{ width:10px; height:10px; border-radius:3px; }}
.legend-dot.green {{ background:#34D399; }}
.legend-dot.yellow {{ background:#FBBF24; }}
.legend-dot.red {{ background:#F87171; }}
.legend-dot.gray {{ background:#475569; }}
.search-bar {{ margin-bottom:12px; display:flex; gap:12px; align-items:center; }}
.search-bar input {{ background:#1E293B; border:1px solid #334155; border-radius:6px; padding:8px 14px; color:#E2E8F0; font-size:13px; width:300px; font-family:'DM Sans',sans-serif; }}
.search-bar input::placeholder {{ color:#475569; }}
.search-bar input:focus {{ outline:none; border-color:#818CF8; }}
.filter-btn {{ background:#1E293B; border:1px solid #334155; border-radius:6px; padding:8px 14px; color:#94A3B8; font-size:12px; cursor:pointer; font-family:'DM Sans',sans-serif; }}
.filter-btn:hover {{ border-color:#818CF8; color:#E2E8F0; }}
.filter-btn.active {{ background:rgba(129,140,248,0.15); border-color:#818CF8; color:#818CF8; }}
.var-count {{ font-size:10px; color:#64748B; margin-left:6px; }}
</style>
</head>
<body>
<div class="container">
<header>
  <h1>VAM Creative Performance Dashboard</h1>
  <p>Value Added Moving — All-Time Performance Data | Generated March 25, 2026</p>
</header>

<div class="legend" id="legend"></div>

<div class="search-bar">
  <input type="text" id="searchInput" placeholder="Search creatives..." oninput="filterTable()">
  <button class="filter-btn" onclick="toggleFilter('winner')" id="filterWinner">Winners Only</button>
  <button class="filter-btn" onclick="toggleFilter('minSpend')" id="filterMinSpend">Min $1,500 Spend</button>
</div>

<div class="tabs" id="tabBar"></div>
<div class="table-wrap" id="tableWrap"></div>

</div>

<script>
const DATA = {{
  static: {static_json},
  video: {video_json},
  copy: {copy_json},
  staticConcepts: {static_concepts_json},
  videoConcepts: {video_concepts_json},
  copyConcepts: {copy_concepts_json}
}};

const TABS = [
  {{ id:'static', label:'Static Ads', data:DATA.static, hasVideo:false, isConcept:false }},
  {{ id:'staticConcepts', label:'Static Concepts', data:DATA.staticConcepts, hasVideo:false, isConcept:true }},
  {{ id:'video', label:'Video Ads', data:DATA.video, hasVideo:true, isConcept:false }},
  {{ id:'videoConcepts', label:'Video Concepts', data:DATA.videoConcepts, hasVideo:true, isConcept:true }},
  {{ id:'copy', label:'Copy', data:DATA.copy, hasVideo:false, isConcept:false }},
  {{ id:'copyConcepts', label:'Copy Concepts', data:DATA.copyConcepts, hasVideo:false, isConcept:true }}
];

let activeTab = 'static';
let sortCol = 'roas';
let sortDir = 'desc';
let filters = {{ winner:false, minSpend:false }};

const KPI = {{
  roas: {{ green:1.8, yellow:1.6 }},
  costConnected: {{ green:300, yellow:400 }},
  cps: {{ green:750, yellow:1000 }},
  connClose: {{ green:40, yellow:30 }},
  hookRate: {{ green:30, yellow:20 }},
  holdRate: {{ green:20, yellow:15 }}
}};

function ccBadge(v, metric) {{
  if (v == null) return 'nd';
  const k = KPI[metric];
  if (!k) return 'nd';
  if (metric === 'roas' || metric === 'hookRate' || metric === 'holdRate' || metric === 'connClose') {{
    if (v >= k.green) return 'wk';
    if (v >= k.yellow) return 'of';
    return 'ak';
  }}
  if (v <= k.green) return 'wk';
  if (v <= k.yellow) return 'of';
  return 'ak';
}}

function statusBadge(s) {{
  if (!s || s === 'nan') return '';
  const sl = s.toLowerCase();
  if (sl.includes('winner')) return '<span class="badge status-winner">Winner</span>';
  if (sl.includes('loser')) return '<span class="badge status-loser">Loser</span>';
  if (sl.includes('test')) return '<span class="badge status-testing">Testing</span>';
  if (sl.includes('fence')) return '<span class="badge status-fence">On the Fence</span>';
  if (sl.includes('pause') || sl.includes('killed')) return '<span class="badge status-other">Paused</span>';
  if (s) return '<span class="badge status-other">' + s + '</span>';
  return '';
}}

function fmtD(v) {{ return v == null ? '—' : (v >= 10 ? '$' + v.toLocaleString('en-US',{{maximumFractionDigits:0}}) : '$' + v.toFixed(2)); }}
function fmtN(v) {{ return v == null ? '—' : v.toLocaleString('en-US',{{maximumFractionDigits:0}}); }}
function fmtP(v) {{ return v == null ? '—' : v.toFixed(1) + '%'; }}
function fmtR(v) {{ return v == null ? '—' : v.toFixed(2) + 'x'; }}

function renderTabs() {{
  const bar = document.getElementById('tabBar');
  bar.innerHTML = TABS.map(t =>
    `<div class="tab ${{t.id===activeTab?'active':''}}" onclick="switchTab('${{t.id}}')">
      ${{t.label}}<span class="count">(${{t.data.length}})</span>
    </div>`
  ).join('');
}}

function renderLegend() {{
  const tab = TABS.find(t => t.id === activeTab);
  let html = `
    <div class="legend-item"><div class="legend-dot green"></div>ROAS ≥ 1.8x | $/Connected ≤ $300 | Close ≥ 40%</div>
    <div class="legend-item"><div class="legend-dot yellow"></div>ROAS 1.6–1.8x | $/Connected $300–$400 | Close 30–40%</div>
    <div class="legend-item"><div class="legend-dot red"></div>ROAS < 1.6x | $/Connected > $400 | Close < 30%</div>
    <div class="legend-item"><div class="legend-dot gray"></div>No Data</div>
  `;
  if (tab.hasVideo) {{
    html += `<div class="legend-item" style="margin-left:16px;border-left:1px solid #334155;padding-left:16px;"><div class="legend-dot green"></div>Hook ≥ 30% | Hold ≥ 20%</div>
    <div class="legend-item"><div class="legend-dot yellow"></div>Hook 20–30% | Hold 15–20%</div>
    <div class="legend-item"><div class="legend-dot red"></div>Hook < 20% | Hold < 15%</div>`;
  }}
  document.getElementById('legend').innerHTML = html;
}}

function getFilteredData(data) {{
  let d = [...data];
  const q = document.getElementById('searchInput').value.toLowerCase();
  if (q) d = d.filter(r => r.name.toLowerCase().includes(q));
  if (filters.winner) d = d.filter(r => r.status && r.status.toLowerCase().includes('winner'));
  if (filters.minSpend) d = d.filter(r => r.cost >= 1500);
  return d;
}}

function renderTable() {{
  const tab = TABS.find(t => t.id === activeTab);
  let data = getFilteredData(tab.data);

  data.sort((a, b) => {{
    let va = a[sortCol], vb = b[sortCol];
    if (va == null && vb == null) return 0;
    if (va == null) return 1;
    if (vb == null) return -1;
    if (sortCol === 'name' || sortCol === 'status') {{
      return sortDir === 'asc' ? String(va).localeCompare(String(vb)) : String(vb).localeCompare(String(va));
    }}
    return sortDir === 'asc' ? va - vb : vb - va;
  }});

  const cols = [
    {{ key:'idx', label:'#', sortable:false }},
    {{ key:'name', label: tab.isConcept ? 'Concept' : 'Name', sortable:true }},
    {{ key:'link', label:'Ad', sortable:false }},
  ];

  if (!tab.isConcept) {{
    cols.push({{ key:'status', label:'Status', sortable:true }});
  }} else {{
    cols.push({{ key:'variations', label:'Ads', sortable:true, fmt:fmtN }});
  }}

  cols.push(
    {{ key:'cost', label:'Spend', sortable:true, fmt:fmtD }},
    {{ key:'leads', label:'Leads', sortable:true, fmt:fmtN }},
    {{ key:'cpl', label:'CPL', sortable:true, fmt:fmtD }},
    {{ key:'newLeads', label:'New Leads', sortable:true, fmt:fmtN }},
    {{ key:'cpnl', label:'CPNL', sortable:true, fmt:fmtD }},
    {{ key:'incoming', label:'Calls In', sortable:true, fmt:fmtN }},
    {{ key:'costIncoming', label:'$/Call In', sortable:true, fmt:fmtD }},
    {{ key:'outbound', label:'Outbound', sortable:true, fmt:fmtN }},
    {{ key:'costOutbound', label:'$/Outbound', sortable:true, fmt:fmtD }},
    {{ key:'inboundConnected', label:'Inb Connected', sortable:true, fmt:fmtN }},
    {{ key:'costInboundConnected', label:'$/Inb Conn', sortable:true, fmt:fmtD }},
    {{ key:'outboundConnected', label:'Outb Connected', sortable:true, fmt:fmtN }},
    {{ key:'costOutboundConnected', label:'$/Outb Conn', sortable:true, fmt:fmtD }},
    {{ key:'connected', label:'Connected', sortable:true, fmt:fmtN }},
    {{ key:'costConnected', label:'$/Connected', sortable:true, fmt:fmtD, badge:'costConnected' }},
  );

  if (tab.hasVideo) {{
    cols.push(
      {{ key:'hookRate', label:'Hook %', sortable:true, fmt:fmtP, badge:'hookRate' }},
      {{ key:'holdRate', label:'Hold %', sortable:true, fmt:fmtP, badge:'holdRate' }}
    );
  }}

  cols.push(
    {{ key:'connClose', label:'Close %', sortable:true, fmt:fmtP, badge:'connClose' }},
    {{ key:'sales', label:'Sales', sortable:true, fmt:fmtN }},
    {{ key:'cps', label:'CAC', sortable:true, fmt:fmtD, badge:'cps' }},
    {{ key:'revenue', label:'Revenue', sortable:true, fmt:fmtD }},
    {{ key:'roas', label:'ROAS', sortable:true, fmt:fmtR, badge:'roas' }},
    {{ key:'profit', label:'Profit', sortable:true, fmt:fmtD }},
    {{ key:'ctrAll', label:'CTR (all)', sortable:true, fmt:fmtP }},
    {{ key:'ctrLink', label:'CTR (link)', sortable:true, fmt:fmtP }},
    {{ key:'cpm', label:'CPM', sortable:true, fmt:fmtD }},
    {{ key:'impressions', label:'Impr', sortable:true, fmt:fmtN }},
  );

  let thead = '<tr>' + cols.map(c => {{
    if (!c.sortable) return `<th>${{c.label}}</th>`;
    const isSorted = sortCol === c.key;
    const arrow = isSorted ? (sortDir === 'asc' ? '▲' : '▼') : '▲';
    return `<th class="${{isSorted?'sorted':''}}" onclick="doSort('${{c.key}}')">
      ${{c.label}}<span class="arrow">${{arrow}}</span>
    </th>`;
  }}).join('') + '</tr>';

  let tbody = data.map((r, i) => {{
    let cells = cols.map(c => {{
      if (c.key === 'idx') return `<td class="idx-col">${{i+1}}</td>`;
      if (c.key === 'name') {{
        const varLabel = r.variations ? `<span class="var-count">(${{r.variations}} ads)</span>` : '';
        return `<td class="name-col" title="${{r.name}}">${{r.name}}${{varLabel}}</td>`;
      }}
      if (c.key === 'link') {{
        if (r.link) {{
          return `<td style="text-align:center"><a href="${{r.link}}" target="_blank" rel="noopener" style="color:#818CF8;text-decoration:none;font-family:'DM Sans',sans-serif;font-size:11px;font-weight:600;">View <svg width="10" height="10" viewBox="0 0 12 12" fill="none" style="vertical-align:-1px;margin-left:2px"><path d="M3.5 1.5H10.5V8.5M10.5 1.5L1.5 10.5" stroke="#818CF8" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/></svg></a></td>`;
        }}
        return '<td style="text-align:center;color:#334155;">—</td>';
      }}
      if (c.key === 'status') return `<td>${{statusBadge(r.status)}}</td>`;

      const v = r[c.key];
      const display = c.fmt ? c.fmt(v) : (v == null ? '—' : v);

      if (c.badge && v != null) {{
        const cls = ccBadge(v, c.badge);
        return `<td class="num-col"><span class="badge ${{cls}}">${{display}}</span></td>`;
      }}

      let style = '';
      if (c.key === 'profit' && v != null) {{
        style = v >= 0 ? 'color:#34D399' : 'color:#F87171';
      }}

      return `<td class="num-col" ${{style?'style="'+style+'"':''}}>${{display}}</td>`;
    }}).join('');
    return `<tr>${{cells}}</tr>`;
  }}).join('');

  document.getElementById('tableWrap').innerHTML = `<table><thead>${{thead}}</thead><tbody>${{tbody}}</tbody></table>`;
}}

function switchTab(id) {{
  activeTab = id;
  sortCol = 'roas';
  sortDir = 'desc';
  renderTabs();
  renderLegend();
  renderTable();
}}

function doSort(col) {{
  if (sortCol === col) {{ sortDir = sortDir === 'asc' ? 'desc' : 'asc'; }}
  else {{ sortCol = col; sortDir = (col === 'name' || col === 'status') ? 'asc' : 'asc'; }}
  renderTable();
}}

function filterTable() {{
  renderTable();
}}

function toggleFilter(f) {{
  filters[f] = !filters[f];
  document.getElementById(f === 'winner' ? 'filterWinner' : 'filterMinSpend').classList.toggle('active');
  renderTable();
}}

switchTab('static');
</script>
</body>
</html>'''

with open('vam_creative_dashboard.html', 'w') as f:
    f.write(dashboard_html)

print(f'Dashboard generated: vam_creative_dashboard.html')
print(f'Static ads: {len(static_data)} -> {len(static_concepts)} concepts')
print(f'Video ads: {len(video_data)} -> {len(video_concepts)} concepts')
print(f'Copy: {len(copy_data)} -> {len(copy_concepts)} concepts')
